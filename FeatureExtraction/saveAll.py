import os, glob
path = 'K:\Old\E_DRIVE\ACADEMICS IV\BTP\April\FE\Pydownloaded'

import openpyxl
wb=openpyxl.load_workbook('saveAllF.xlsx')
ws=wb.active

##### Import user defiined  Functions ###
#from clrAvg import ImgAnalyze
#from FeatureVector import getFeatureVector
#from IlluminationValues import getIlValues

from FeatureVector1 import getFeatureVector

## Use Defined Functions - A == 0,AA==26,AB== 27
def colToNum(col_str):
    """ Convert base26 column string to number. """
    expn = 0
    col_num = 0
    for char in reversed(col_str):
        col_num += (ord(char) - ord('A') + 1) * (26 ** expn)
        expn += 1

    return (col_num-1)



print("PROCESSING ...")
counter=0;          ## Number of Files processed
rowNumber=0;        ## A1 == (row==0 && column ==0)

for infile in glob.glob( os.path.join(path, '*.*') ):
    #####  FEATURE EXTRACTION AND DATA RETRIEVAL    #####
    ###############
    #if counter == 4 :
    #    break;
    ##############
    base=os.path.basename(infile);    imgID=base[:-9];

    #### get all ####
    fv=getFeatureVector(base)      #base=str(filename.ext)

    
    if fv[0] == -2 :
             print("FileNo: %d"%(counter+1),end="-")
             print("DELETING INVALID FILE: "+base)
             os.remove(path+"\\"+base)   ##delete INVALID files and CONTINUE  #counter+=1; #removing File---> No need of row
             continue;

    
    rowNumber = counter+2  ; 
	#initial one is left for column names
	#if(row number starts with 0){initialise rowNumber=counter+1}
	#elif(row number starts with 1){go with initialsed already.}

    #####     DATA STORING      #####
    #imgID
    ws.cell(row = rowNumber, column = 1 ).value=imgID  ;  #Column 1 is for imgID
   			
	#Combined			
    colFirst=2
    colLast=(colFirst + 28)
				
    j=0;
    for col in range(colFirst,colLast):
                ws.cell(row = rowNumber, column = col).value = fv[j]
                j+=1
	
				

    if counter%5 == 0  :
        print(counter,end='-')
      
    
    counter+=1;
#####  END OF PROCESSING   #####

# Saving the Changes in Excel File
wb.save('saveAllF.xlsx')

############# END ###############
print("PROCESSED ALL FILES SUCCESSFULLY!")
print(counter)
