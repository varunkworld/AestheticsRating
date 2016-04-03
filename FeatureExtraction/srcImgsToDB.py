import os, glob
path = 'K:\Old\E_DRIVE\ACADEMICS IV\BTP\April\FE\Pydownloaded'

import openpyxl
wb=openpyxl.load_workbook('FE_data.xlsx')
ws=wb.active

##### Import user defiined  Functions ###
from clrAvg import ImgAnalyze
from FeatureVector import getFeatureVector


print("PROCESSING ...")
counter=0;          ## Number of Files processed
rowNumber=0;        ## A1 == (row==0 && column ==0)

for infile in glob.glob( os.path.join(path, '*.*') ):
    #####  FEATURE EXTRACTION AND DATA RETRIEVAL    #####

    base=os.path.basename(infile);    imgID=base[:-9];

    #### RGB AVGs ####
    rAvg,gAvg,bAvg=ImgAnalyze(base)      #base=str(filename.ext)
    if rAvg == -2 :
             print("FileNo: %d"%(counter+1),end="-")
             print("DELETING INVALID FILE: "+base)
             os.remove(path+"\\"+base)   ##delete INVALID files and CONTINUE  #counter+=1; #removing File---> No need of row
             continue;


            
                            ## Get Feature Vector ##
    fv= getFeatureVector(base);


    rowNumber = counter+2  ;

    #####     DATA STORING      #####
    #imgID
    ws.cell(row = rowNumber, column = 0 ).value=imgID  ;  #Column 0 is for imgID
    #RGBAVG
    ws.cell(row = rowNumber, column = 1 ).value=rAvg      
    ws.cell(row = rowNumber, column = 2 ).value=gAvg
    ws.cell(row = rowNumber, column = 3 ).value=bAvg
    #FV
    ws.cell(row = rowNumber, column = 4 ).value=fv[0]; 
    ws.cell(row = rowNumber, column = 5 ).value=fv[1];
    ws.cell(row = rowNumber, column = 6 ).value=fv[2];
    ws.cell(row = rowNumber, column = 7 ).value=fv[3];
    ws.cell(row = rowNumber, column = 8 ).value=fv[4];
    ws.cell(row = rowNumber, column = 9 ).value=fv[5];
    ws.cell(row = rowNumber, column = 10 ).value=fv[6];
    ws.cell(row = rowNumber, column = 11 ).value=fv[7];

    #initial two are left for column names
    counter+=1;
#####  END OF PROCESSING   #####

# Saving the Changes in Excel File
wb.save('FE_data.xlsx')

############# END ###############
print("PROCESSED ALL FILES SUCCESSFULLY!")
print(counter)
